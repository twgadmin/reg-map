from ensurepip import version
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.generic import ListView, CreateView, UpdateView
from django.views import View
from rest_framework import viewsets
from django.http import FileResponse, JsonResponse
from django.db.models import Prefetch
import re
from apps.common.mixins import RequestInjectMixin
from apps.project.forms import ProjectForm, VersionForm, RegisterForm
from apps.project.models import Project, Version, Register, Field, FieldChoice
from apps.project.serializers import RegisterSerializer, FieldSerializer, FieldChoiceSerializer, VersionSerializer
import xlsxwriter
import openpyxl
from rest_framework.response import Response

class CommonMessageMixin(SuccessMessageMixin):
    """
    Since all of the forms are pop-up, showing message on field is difficult
    So message are shown as error messages
    """

    def form_invalid(self, form):
        # Extract Error from forms
        for k, v in form.errors.items():
            messages.error(self.request, v)
        return redirect(self.request.META["HTTP_REFERER"])


class CommonVersionCreateUpdateMixin(
    CommonMessageMixin,
    RequestInjectMixin,
    LoginRequiredMixin,
):
    """
    Copy common logic of create and update implementation if Version Crud
    """

    queryset = Version.objects.all()
    form_class = VersionForm
    success_url = reverse_lazy("project:version-list")
    template_name = "project/snippet/form-create-update.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()

        # Disabled project field while creating project from version list page
        kwargs["project_field_disabled"] = bool(
            self.request.GET.get("project_field_disabled")
        )
        return kwargs

    def get_success_url(self):
        """Return the URL to redirect to after processing a valid form."""

        # If next_url is passed from frontend redirect to that url after success
        next_url = self.request.GET.get("next")
        if next_url and url_has_allowed_host_and_scheme(next_url, settings.ALLOWED_HOSTS):
            return redirect(next_url)

        return reverse_lazy(
            "project:version-list", kwargs={"pk": self.object.project.id}
        )

    def get_initial(self):
        initial = super().get_initial()
        # If project_id is passed as query parameters set that initial value
        project_id = self.request.GET.get("project")
        if project_id:
            initial["project"] = get_object_or_404(Project, id=project_id)
        return initial


class CommonRegisterCreateUpdateMixin(
    CommonMessageMixin,
    RequestInjectMixin,
    LoginRequiredMixin,
):
    """
    Copy common logic of create and update implementation if Register Crud
    """

    queryset = Register.objects.all()
    form_class = RegisterForm
    success_url = reverse_lazy("project:register-list")
    template_name = "project/snippet/form-create-update.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()

        # Disabled project field while creating project from version list page
        kwargs["version_field_disabled"] = bool(
            self.request.GET.get("version_field_disabled")
        )
        return kwargs

    def get_success_url(self):
        """Return the URL to redirect to after processing a valid form."""
        return reverse_lazy(
            "project:register-list", kwargs={"pk": self.object.version.id}
        )


class ProjectCreateView(
    RequestInjectMixin, CommonMessageMixin, LoginRequiredMixin, CreateView
):
    queryset = Register.objects.all()
    serializer_class = RegisterSerializer

    def post(self, request, *args, **kwargs):
        return self.create(request, *args, **kwargs)

    success_message = "Successfully Created New Project"


class ProjectUpdateView(
    RequestInjectMixin, CommonMessageMixin, LoginRequiredMixin, UpdateView
):

    queryset = Project.objects.all()
    form_class = ProjectForm
    success_url = reverse_lazy("common:home")
    template_name = "project/snippet/form-create-update.html"
    success_message = "Successfully Updated Project"


class VersionListView(LoginRequiredMixin, ListView):

    queryset = Version.objects.all()
    template_name = "project/version-list.html"
    context_object_name = "versions"
    paginate_by = 5

    def get_queryset(self):
        """
        Filter versions using project id provided in URL
        """
        return super().get_queryset().filter(project_id=self.kwargs.get("pk"))

    def get_context_data(self, *args, **kwargs):
        version_id = self.request.GET.get("register_id")
        status = self.request.GET.get("status")
        type = self.request.GET.get("type")

        if status == "V":
            if type == "delete":
                version = Version.objects.get(id=version_id)
                version.delete()
                return
        else:
            data = super().get_context_data(*args, **kwargs)
            data["project"] = get_object_or_404(
                Project, id=self.kwargs.get("pk"))
            return data


class VersionCreateView(CommonVersionCreateUpdateMixin, CreateView):

    success_message = "Successfully Created New Version"


class VersionUpdateView(CommonVersionCreateUpdateMixin, UpdateView):

    success_message = "Successfully Updated Version"


class VersionCopyView(View):
    def get(self, request, *args, **kwargs):
        version_id = self.request.GET.get("version")
        if version_id:
            old_version = Version.objects.get(id=version_id)
            old_register = Register.objects.filter(version_id=old_version.id)
            new_version = Version.objects.create(
                project_id=old_version.project_id, name="Copy "+old_version.name)
            for r in old_register:
                new_register = Register.objects.create(version_id=new_version.id, name=r.name, address=r.address,
                                                       size=r.size, description=r.description)
                old_field = Field.objects.filter(register_id=r.id)
                for f in old_field:
                    new_field = Field.objects.create(register_id=new_register.id, range_upper=f.range_upper, range_lower=f.range_lower,
                                                     type=f.type, unit_scale=f.unit_scale, unit_type=f.unit_type, name=f.name, description=f.description,
                                                     value_min=f.value_min, value_max=f.value_max)
                    old_choice = FieldChoice.objects.filter(field_id=f.id)
                    for c in old_choice:
                        FieldChoice.objects.create(
                            field_id=new_field.id, name=c.name, desc=c.desc, value=c.value)
            return redirect("project:version-list", old_version.project_id)


class FieldListView(LoginRequiredMixin, ListView):
    queryset = Field.objects.all()
    template_name = "project/field-list.html"
    context_object_name = "fields"
    paginate_by: int = 500

    def get_queryset(self):
        return list(super().get_queryset())

    def get_context_data(self, *args, **kwargs):
        register_id = self.request.GET.get("register_id")
        field_id = self.request.GET.get("field_id")
        choice_id = self.request.GET.get("choice_id")
        status = self.request.GET.get("status")
        type = self.request.GET.get("type")
        value = self.request.GET.get("value")

        if status == "F":
            if type == "delete":
                f = Field.objects.get(id=register_id)
                f.delete()
                return
            else:
                f = Register.objects.get(
                    id=register_id).fields.get(id=field_id)
                setattr(f, type, value)
                f.save()
        elif status == "C":
            if type == "delete":
                choice = FieldChoice.objects.get(id=register_id)
                choice.delete()
            else:
                choice = (
                    Register.objects.get(id=register_id)
                    .fields.get(id=field_id)
                    .choices.get(id=choice_id)
                )
                setattr(choice, type, value)
                choice.save()
        else:
            data = super().get_context_data(*args, **kwargs)
            return data


class RegisterListView(LoginRequiredMixin, ListView):
    queryset = Register.objects.all()
    template_name = "project/register-list.html"
    context_object_name = "registers"
    paginate_by = 500

    def get_queryset(self):
        """
        Filter versions using version id provided in URL
        """
        registers = (
            super()
            .get_queryset()
            .filter(version_id=self.kwargs.get("pk"))
            .prefetch_related('fields')
            .order_by("address")
        )
        return list(registers)

    def get_context_data(self, *args, **kwargs):

        register_id = self.request.GET.get("register_id")
        field_id = self.request.GET.get("field_id")
        choice_id = self.request.GET.get("choice_id")
        status = self.request.GET.get("status")
        type = self.request.GET.get("type")
        value = self.request.GET.get("value")
        if status == "R":
            r = Register.objects.get(id=register_id)
            if type == "address":
                value = int(value, 16)
            if type == "delete":
                r.delete()
                return
            setattr(r, type, value)
            r.save()

        elif status == "F":
            if type == "delete":
                f = Field.objects.get(id=register_id)
                f.delete()
                return
            else:
                f = Register.objects.get(
                    id=register_id).fields.get(id=field_id)
                setattr(f, type, value)
                f.save()
        elif status == "C":
            if type == "delete":
                choice = FieldChoice.objects.get(id=register_id)
                choice.delete()
            else:
                choice = (
                    Register.objects.get(id=register_id)
                    .fields.get(id=field_id)
                    .choices.get(id=choice_id)
                )
                setattr(choice, type, value)
                choice.save()
        elif status == "N":
            if type == "reserved":
                addr = int(value[0:6], 16)
                size = int(value[-6: len(value)], 16) - int(value[0:6], 16)
                Register.objects.create(
                    version_id=register_id, name="dummy", address=addr, size=size
                )
            else:
                if field_id == None:
                    f = Field.objects.create(
                        register_id=register_id,
                        range_upper=15,
                        range_lower=0,
                        value_max=0,
                        value_min=0,
                        name="dummy",
                    )
                    setattr(f, type, value)
                    f.save()
                else:
                    f = FieldChoice.objects.create(field_id=field_id)
                    setattr(f, type, value)
                    f.save()
        data = super().get_context_data(*args, **kwargs)
        data["version"] = get_object_or_404(
            Version.objects
            .get_queryset()
            .prefetch_related(
                Prefetch(
                    "registers",
                    queryset=Register.objects.prefetch_related(
                        Prefetch(
                            # Depending how these are related you may be able to
                            # save some queries using `select_related()` instead.
                            "fields",
                            queryset=Field.objects.prefetch_related(
                                Prefetch(
                                    "choices",
                                    queryset=FieldChoice.objects.order_by(
                                        "value")
                                )
                            ).order_by('range_upper')
                        )
                    )
                )
            ), id=self.kwargs.get("pk"))
        data["project"] = data["version"].project
        return data


class RegisterCreateView(CommonRegisterCreateUpdateMixin, CreateView):

    success_message = "Successfully Created New Register"


class RegisterUpdateView(CommonRegisterCreateUpdateMixin, UpdateView):

    success_message = "Successfully Updated Register"


class VersionViewSet(viewsets.ModelViewSet):
    queryset = Version.objects.all()
    serializer_class = VersionSerializer


class RegisterViewSet(viewsets.ModelViewSet):
    queryset = Register.objects.all()
    serializer_class = RegisterSerializer

    def list(self, request, *args, **kwargs):
        queryset=Register.objects.prefetch_related(
                        Prefetch(
                            # Depending how these are related you may be able to
                            # save some queries using `select_related()` instead.
                            "fields",
                            queryset=Field.objects.prefetch_related(
                                Prefetch(
                                    "choices",
                                    queryset=FieldChoice.objects.order_by(
                                        "value")
                                )
                            ).order_by('-range_upper')
                        )
                    )
        # queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


    def create(self, request, *args, **kwargs):
        new_register = Register.objects.create(
            version_id = 1,
            address = int(request.data['address'][0:6], 16),
            name = 'New Entry',
            size = int(request.data['address'][-6: len(request.data['address'])], 16) - int(request.data['address'][0:6], 16)
        )
        new_register = RegisterSerializer(new_register)
        return JsonResponse({"new_record": new_register.data, "status": "new record success"})


class FieldViewSet(viewsets.ModelViewSet):
    queryset = Field.objects.all()
    serializer_class = FieldSerializer

    def create(self, request, *args, **kwargs):
        new_field = Field.objects.create(
            range_upper = request.data['range_upper'],
            range_lower = request.data['range_lower'],
            register_id = request.data['register_id'],
        )
        new_record = FieldSerializer(new_field)
        return JsonResponse({"new_record": new_record.data, "status": "new record success"})

    def update(self, request, *args, **kwargs):
        overlapping_records_range_lower = Field.objects.filter(range_lower__lte=request.data['range_lower'], range_upper__gte=request.data['range_lower'], register_id=request.data['register_id']).exclude(id__in=[request.data['id']])
        overlapping_records_range_upper = Field.objects.filter(range_lower__lte=request.data['range_upper'], range_upper__gte=request.data['range_upper'], register_id=request.data['register_id']).exclude(id__in=[request.data['id']])
        if len(overlapping_records_range_lower or overlapping_records_range_upper) > 0:
            return JsonResponse({"status": "field update failure", "message": "Invalid input for value_min and value_max"}, status=400) 

        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, '_prefetched_objects_cache', None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return JsonResponse({"status": "field update success", "message": "Updated field successfully."}, status=200)

    def perform_update(self, serializer):
        serializer.save()


class FieldChoiceViewSet(viewsets.ModelViewSet):
    queryset = FieldChoice.objects.all()
    serializer_class = FieldChoiceSerializer

    def create(self, request, *args, **kwargs):
        new_field_choice = FieldChoice.objects.create(
            value = request.data['value'],
            name = 'New Entry',
            field_id = request.data['field_id'],
        )
        new_record = FieldChoiceSerializer(new_field_choice)
        return JsonResponse({"new_record": new_record.data, "status": "new record success"})


def importFunc(file, version_id):
    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook("./upload/" + file)
    # Define variable to read the active sheet:
    worksheet = wookbook.active
    # Iterate the loop to read the cell values
    # Iterate the loop to read the cell values
    for i in range(1, worksheet.max_row):
        index = 0
        new_register = {}
        fields = list()
        for col in worksheet.iter_cols(1, worksheet.max_column):
            index += 1
            if index == 1:
                new_register["address"] = int(str(col[i].value), 16)
            elif index == 2:
                new_register["name"] = col[i].value
            elif index == 3:
                new_register["access"] = str(col[i].value)
            elif index == 4:
                new_register["size"] = int(col[i].value)
            elif index == 5:
                new_register["default"] = str(col[i].value)
            else:
                try:
                    description = re.findall(
                        r'[\s\S]*?(?=[0-9]+:[0-9]+)', str(col[i].value), re.DOTALL)[0]
                except:
                    description = str(col[i].value)
                new_register["description"] = description
                data = str(col[i].value).split("\n")

                for line in data:
                    line_is_field = re.match("^[0-9]+:[0-9]+", line)
                    if line_is_field:
                        if "reserved" in line.lower() or \
                                "RxDxRVxD".lower() in line.lower():
                            continue  # skip this line
                        field_params = line.split("-")
                        field_params = [s.strip() for s in field_params]
                        upper, lower = field_params[0].split(":")
                        try:
                            field_type = field_params[1]
                        except:
                            pass
                        field_range = field_params[2].strip(
                            '[]') if 2 < len(field_params) else None
                        scale_unit = field_params[3].strip(
                            '[]') if 3 < len(field_params) else None
                        name = field_params[4] if 4 < len(field_params) else ""
                        desc = field_params[5] if 5 < len(field_params) else ""

                        if isinstance(scale_unit, str):
                            try:
                                scale, unit = scale_unit.split(",")
                            except:
                                scale, unit = None, None

                        if field_range is not None and \
                           field_range != "None" and \
                           field_range != '':
                            try:
                                value_min, value_max = field_range.split(",")
                            except:
                                value_min, value_max = None, None
                        else:
                            value_min, value_max = None, None
                        f = Field(range_upper=upper, range_lower=lower, type=field_type, unit_scale=scale,
                                  unit_type=unit, name=name, description=desc, value_min=value_min, value_max=value_max)
                        fields.append(f)
        r, created = Register.objects.get_or_create(
            version_id=version_id,
            address=new_register["address"],
            defaults={
                'name': new_register["name"],
                'size': new_register["size"],
                'description': new_register["description"]
            }
        )
        Field.objects.filter(register=r).delete()
        for f in fields:
            f.register = r
            f.save()


class FileUploadView(View):
    # parser_classes = (FileUploadParser,)
    def post(self, request,  format='xlsx'):
        up_file = request.FILES['file']
        version_id = request.POST["version"]
        destination = open('./upload/' + up_file.name, 'wb+')
        for chunk in up_file.chunks():
            destination.write(chunk)
        destination.close()

        importFunc(up_file.name, version_id)

        return redirect("project:register-list", version_id)


class FileDownloadView(View):
    def get(self, *args, **kwargs):
        version_id = self.kwargs.get("pk")
        project_name = Version.objects.get(id=version_id).project.name
        version_name = Version.objects.get(id=version_id).name
        registers = Register.objects.filter(version_id=version_id)
        filename = project_name + "_" + version_name + ".xlsx"

        workbook = xlsxwriter.Workbook('./download'+filename)
        worksheet = workbook.add_worksheet()
        col = 0
        row = 0
        # set the header
        worksheet.write(row, col, "Address_hex")
        worksheet.write(row, col + 1, "name")
        worksheet.write(row, col + 2, "access")
        worksheet.write(row, col + 3, "address_length")
        worksheet.write(row, col + 4, "default")
        worksheet.write(row, col + 5, "description")
        row += 1
        for r in registers:
            worksheet.write(row, col, hex(r.address))
            worksheet.write(row, col + 1, r.name)
            worksheet.write(row, col + 2, r.access)
            worksheet.write(row, col + 3, r.size)
            worksheet.write(row, col + 4, r.default)
            desc = r.description + "\n"
            for f in r.fields.all():
                desc = (
                    desc
                    + str(f.range_upper)
                    + ":"
                    + str(f.range_lower)
                    + " - "
                    + str(f.type)
                    + " - "
                    + str(f.unit_scale)
                    + " - "
                    + str(f.description)
                    + "\n"
                )
                for c in f.choices.all():
                    desc = (
                        desc
                        + "   "
                        + str(c.value)
                        + " - "
                        + str(c.name)
                        + " - "
                        + str(c.desc)
                        + "\n"
                    )
            worksheet.write(row, col + 5, desc)
            row += 1
        return FileResponse(open("./download/"+filename, 'rb'))


class UpdateDataView(View):
    def get(self, *args, **kwargs):
        version_id = self.request.GET.get('version_id')
        register_id = self.request.GET.get("register_id")
        field_id = self.request.GET.get("field_id")
        choice_id = self.request.GET.get("choice_id")
        status = self.request.GET.get("status")
        type = self.request.GET.get("type")
        value = self.request.GET.get("value")
        new_record = None
        if status == "R":
            r = Register.objects.get(id=register_id)
            if type == "address":
                value = int(value, 16)
            if type == "delete":
                r.delete()
                return JsonResponse({"status": "delete success"})
            setattr(r, type, value)
            r.save()

        elif status == "F":
            if type == "delete":
                f = Field.objects.get(id=register_id)
                f.delete()
                return JsonResponse({"status": "delete success"})
            else:
                f = Register.objects.get(
                    id=register_id).fields.get(id=field_id)
                if type == "value_min":
                    if "," in value:
                        values = value.split(",")
                        setattr(f, "value_min", values[0])
                        setattr(f, "value_max", values[1])
                    else:
                        setattr(f, "value_min", value)
                        setattr(f, "value_max", None)
                else:
                    setattr(f, type, value)
                f.save()
        elif status == "C":
            if type == "delete":
                choice = FieldChoice.objects.get(id=register_id)
                choice.delete()
                return JsonResponse({"status": "delete success"})
            else:
                choice = (
                    Register.objects.get(id=register_id)
                    .fields.get(id=field_id)
                    .choices.get(id=choice_id)
                )
                setattr(choice, type, value)
                choice.save()
        elif status == "N":
            if type == "reserved":
                addr = int(value[0:6], 16)
                size = int(value[-6: len(value)], 16) - int(value[0:6], 16)
                new_record = Register.objects.create(
                    version_id=version_id, name="New Entry", address=addr, size=size
                )
                new_record = RegisterSerializer(new_record)
            else:
                if field_id == None:
                    f = Field.objects.create(
                        register_id=register_id,
                        range_upper=15,
                        range_lower=0,
                        value_max=0,
                        value_min=0,
                        name="New Entry",
                    )
                    setattr(f, type, value)
                    f.save()
                    new_record = FieldSerializer(f)
                else:
                    f = FieldChoice.objects.create(field_id=field_id)
                    setattr(f, type, value)
                    f.save()
                    new_record = FieldChoiceSerializer(f)
        if new_record:
            return JsonResponse({"new_record": new_record.data, "status": "new record success"})
        else:
            return JsonResponse({"status": "update success"})
